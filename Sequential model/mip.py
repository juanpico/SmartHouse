# %%
import gurobipy as gp
import numpy as np
import pandas as pd
import openpyxl as op

# %%
df = pd.read_excel("../Data/instances/test_instance.xlsx", sheet_name="series")
pm_data = pd.read_excel("../Data/instances/test_instance.xlsx", sheet_name="PM")

# %%

# 1. Conjuntos

# Conjunto de timesteps
T = np.arange(df.shape[0])
# Conjunto de appliances
N = np.arange(pm_data.shape[1]-1)

#%%
# 2. Parámetros

# g_t: Energy genereated from PV panels at time t
g = df['solar'].to_numpy()
# l_t: Unshiftable load at time t
l = df['loads'].to_numpy()
# v_n: load of appliance n
v = np.array([10, 15, 20, 5, 10])
# r_n: runtime of appliance n
r = np.array([4, 3, 2, 2, 3])
# pm_(n.t): User preference of using appliance n at time t
pm = pm_data.drop('time', axis=1).to_numpy().transpose()
# p_t: Buying price of electricity from the grid at time t
p = df['price'].to_numpy()
# gamma_t: selling price of electricity from the grid at time t
gamma = p*0.8
# eta: charge and discharge efficiency of the battery
eta = 1
# e_max: Capacity of the battery
e_max = 200
# e_min: Minimun energy stored in the battery
e_min = 0
# e_0: Initial energy stored in the battery
e_0 =  0
# f_min: lower bound of discharge rate
f_min = -25
# f_max: upper bound of charge rate
f_max = 25 
# delta: size of timestep (in hours)
delta = 15/60
# alpha: objetive hyperparameter
alpha = 0.01

# %%

########################## MODELO DE OPTIMIZACIÓN ############################

# 3. Definir variables de decisión
model = gp.Model("smart_house")

# b_t: energy bought from the grid at time t
b = model.addVars(T, vtype = gp.GRB.CONTINUOUS , name = 'b')

# s_t: energy sold to the grid at time t
s = model.addVars(T, vtype = gp.GRB.CONTINUOUS, name = 's')

# a_(n,t): 1 if appliance n opereates at time t
a = model.addVars(N,T, vtype = gp.GRB.BINARY, name = 'a')

# w_(n,t): 1 if appliance n starts to operate at time t
w = model.addVars(N,T, vtype = gp.GRB.BINARY, name = 'w')

# f_t: energy transferred to or from the battery at time t
f =  model.addVars(T, lb=float('-inf'), vtype = gp.GRB.CONTINUOUS, name = 'f')

# e_t: energy stored in the battery at time t
e = model.addVars(T, vtype = gp.GRB.CONTINUOUS, name = 'e')

# d_t: user discomfort at t
d = model.addVars(T, vtype = gp.GRB.CONTINUOUS, name = 'd')

# c_t: electricity costs at t
c = model.addVars(T, vtype = gp.GRB.CONTINUOUS, name = 'c')


# 4. Restricciones

# 4.1. Power balance
for t in T:
   model.addConstr(b[t]+g[t]==f[t]+l[t]+gp.quicksum(delta*v[n]*a[n,t] for n in N)+s[t])
   

# 4.2. Battery dynamics
model.addConstr(e[0]==e_0+eta*f[0])
for t in T:
    if t>0:
        model.addConstr(e[t]==e[t-1]+eta*f[t])
        

# 4.3. Battery capacity and charging bounds
for t in T:
    model.addConstr(e[t]>=e_min)
    model.addConstr(e[t]<=e_max)
    model.addConstr(f[t]>=f_min)
    model.addConstr(f[t]<=f_max)

# 4.4. Shiftable appliances

# Runtime
#for n in N:
#    model.addConstr(gp.quicksum(a[n,t] for t in T) == r[n])

# Consecutive runtime
for n in N:
    for t in T:
       # if t <= T.shape[0]-r[n]+1 -1:         
        model.addConstr(gp.quicksum(a[n,k] for k in np.arange(t, min(t+r[n]-1 + 1, T.shape[0]-1)))>=r[n]*w[n,t])

#for n in N:
#    model.addConstr(gp.quicksum(w[n,t] for t in T if t <= T.shape[0]-r[n]+1 -1) == 1)

for n in N:
    model.addConstr(gp.quicksum(w[n,t] for t in T) == 1)


# 4.5. User discomfort
for t in T:
    model.addConstr(d[t] == gp.quicksum(a[n,t]*(1-pm[n,t]) for n in N))

# 4.6. Electricity costs
for t in T:
    model.addConstr(c[t] == b[t]*p[t] - gamma[t]*s[t])

# %%

# 5. Función Objetivo

model.setObjective(gp.quicksum(alpha*c[t]+(1-alpha)*d[t] for t in T), gp.GRB.MINIMIZE)

model.update()
model.optimize()

# %%

# 6. Resultados

book = op.Workbook()

sheet = book.active
sheet.title = 'Resultados'
sheet.cell(1,1).value = 'F.O.'
sheet.cell(1,2).value = model.ObjVal
sheet.cell(3,1).value = 'time'
sheet.cell(3,2).value = 'b'
sheet.cell(3,3).value = 's'
sheet.cell(3,4).value = 'f'
sheet.cell(3,5).value = 'e'
sheet.cell(3,6).value = 'd'
sheet.cell(3,7).value = 'c'

for t in T:
    sheet.cell(t+4, 1).value = t
    sheet.cell(t+4, 2).value = b[t].x
    sheet.cell(t+4, 3).value = s[t].x
    sheet.cell(t+4, 4).value = f[t].x
    sheet.cell(t+4, 5).value = e[t].x
    sheet.cell(t+4, 6).value = d[t].x
    sheet.cell(t+4, 7).value = c[t].x

book.create_sheet('appliances a')
sheet = book['appliances a']

sheet.cell(1, 1).value = "time"
for n in N:
    sheet.cell(1, 2+n).value = n

for t in T:
    sheet.cell(t+2, 1).value = t
    for n in N:
        sheet.cell(t+2, 2+n).value = a[n,t].x

book.create_sheet('appliances w')
sheet = book['appliances w']

sheet.cell(1, 1).value = "time"
for n in N:
    sheet.cell(1, 2+n).value = n

for t in T:
    sheet.cell(t+2, 1).value = t
    for n in N:
        sheet.cell(t+2, 2+n).value = w[n,t].x

 
book.save('resultados_test2.xlsx')
# %%
